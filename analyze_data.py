import openpyxl
from openpyxl.utils import get_column_letter

# Count data rows and understand P/F/A grouping in Z1F
print("=== Z1F - Data pattern analysis ===")
wb = openpyxl.load_workbook(r'Attachment 1-Procurement Plan-Z1F.xlsx', data_only=True)
ws = wb['PSR']
pfa_count = {'P': 0, 'F': 0, 'A': 0}
packages = set()
for row_idx in range(13, ws.max_row + 1):
    m_val = ws.cell(row=row_idx, column=13).value  # Col M
    d_val = ws.cell(row=row_idx, column=4).value   # Col D (Package Name)
    a_val = ws.cell(row=row_idx, column=1).value    # Col A (No.)
    if m_val in ('P', 'F', 'A'):
        pfa_count[m_val] += 1
    if a_val is not None and d_val is not None:
        packages.add(d_val)

print(f"P/F/A counts: {pfa_count}")
print(f"Unique packages: {len(packages)}")
print(f"Total data rows: {ws.max_row - 12}")

# Count data rows in ASK
print("\n=== ASK - Data pattern analysis ===")
wb2 = openpyxl.load_workbook(r'Attachment 2-Procurement Plan-ASK.xlsx', data_only=True)
ws2 = wb2['PSR Overall Cycle']
pfa_count2 = {'Plan': 0, 'Forecast': 0, 'Actual': 0}
packages2 = set()
for row_idx in range(19, ws2.max_row + 1):
    p_val = ws2.cell(row=row_idx, column=16).value  # Col P
    c_val = ws2.cell(row=row_idx, column=3).value   # Col C (Package Name)
    a_val = ws2.cell(row=row_idx, column=1).value    # Col A (No.)
    if p_val in ('Plan', 'Forecast', 'Actual'):
        pfa_count2[p_val] += 1
    if a_val is not None and c_val is not None:
        packages2.add(c_val)

print(f"Plan/Forecast/Actual counts: {pfa_count2}")
print(f"Unique packages: {len(packages2)}")
print(f"Total data rows: {ws2.max_row - 18}")

# Show some package names
print("\n=== Z1F Package names (first 10) ===")
for i, p in enumerate(sorted(packages)[:10]):
    print(f"  {i+1}. {p}")

print("\n=== ASK Package names (first 10) ===")
for i, p in enumerate(sorted(packages2)[:10]):
    print(f"  {i+1}. {p}")

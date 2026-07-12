"""
Procurement Tracking Dashboard Server
Reads Excel files (Z1F & ASK) and serves a web dashboard with delay tracking.
Supports dynamic file replacement - just drop new Excel files and click Update.
"""

import http.server
import socketserver
import json
import os
import datetime
import threading
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter

# ─── Configuration ───────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
PORT = int(os.environ.get("PORT", 8080))

# ─── Z1F Column Mapping ─────────────────────────────────────────────────────
# PSR sheet: Header row 10, Data starts row 13, P/F/A indicator in col M
Z1F_CONFIG = {
    "sheet_name": "PSR",
    "header_row": 10,
    "data_start_row": 13,
    "pfa_col": "M",          # P/F/A indicator
    "pfa_values": {
        "P": "Plan", "PLAN": "Plan",
        "F": "Forecast", "FORECAST": "Forecast",
        "A": "Actual", "ACTUAL": "Actual"
    },
    "package_name_col": "D",
    "rfq_no_col": "C",
    "mr_no_col": "B",
    "item_no_col": "A",
    "priority_col": "K",
    "lli_col": "I",
    "stages": [
        {"name": "Bidder List Approval",       "col": "N"},
        {"name": "MR IFA Issued",              "col": "O"},
        {"name": "MR IFA Approved",            "col": None},
        {"name": "MR Issued",                  "col": None},
        {"name": "MR Approved",                "col": "P"},
        {"name": "RFQ Issued",                 "col": "Q"},
        {"name": "RFQ Approved",               "col": "R"},
        {"name": "CFT Issuance",               "col": None},
        {"name": "Bid Closing Date",           "col": "S"},
        {"name": "TBE Issued",                 "col": "T"},
        {"name": "TBE Approved",               "col": "V"},
        {"name": "PO Issued",                  "col": "W"},
        {"name": "Vendor Acknowledgement",     "col": "X"},
        {"name": "KOM",                        "col": "Y"},
        {"name": "VD Submission",              "col": "Z"},
        {"name": "Approval of Key VD",         "col": "AA"},
        {"name": "Delivery of Major Materials","col": "AB"},
        {"name": "PIM",                        "col": "AC"},
        {"name": "Start Production",           "col": "AD"},
        {"name": "FAT",                        "col": "AE"},
        {"name": "Ready for Shipment",         "col": "AF"},
        {"name": "Receipt at Worksite",        "col": "AG"},
        {"name": "Punch List Clearance",       "col": "AH"},
        {"name": "Final Documentation",        "col": "AI"},
    ],
}

# ─── ASK Column Mapping ─────────────────────────────────────────────────────
# PSR Overall Cycle sheet: Header row 16, Data starts row 19, Plan/Forecast/Actual in col P
ASK_CONFIG = {
    "sheet_name": "PSR Overall Cycle",
    "header_row": 16,
    "data_start_row": 19,
    "pfa_col": "P",          # Plan/Forecast/Actual indicator
    "pfa_values": {
        "P": "Plan", "PLAN": "Plan",
        "F": "Forecast", "FORECAST": "Forecast",
        "A": "Actual", "ACTUAL": "Actual"
    },
    "package_name_col": "C",
    "rfq_no_col": "B",
    "mr_no_col": None,
    "item_no_col": "A",
    "priority_col": "I",
    "lli_col": "G",
    "stages": [
        {"name": "Bidder List Approval",       "col": "Q"},
        {"name": "MR IFA Issued",              "col": "R"},
        {"name": "MR IFA Approved",            "col": "S"},
        {"name": "MR Issued",                  "col": "T"},
        {"name": "MR Approved",                "col": "U"},
        {"name": "RFQ Issued",                 "col": "V"},
        {"name": "RFQ Approved",               "col": "W"},
        {"name": "CFT Issuance",               "col": "X"},
        {"name": "Bid Closing Date",           "col": "Y"},
        {"name": "TBE Issued",                 "col": "Z"},
        {"name": "TBE Approved",               "col": "AA"},
        {"name": "PO Issued",                  "col": "AB"},
        {"name": "Vendor Acknowledgement",     "col": "AC"},
        {"name": "KOM",                        "col": "AD"},
        {"name": "VD Submission",              "col": "AE"},
        {"name": "Approval of Key VD",         "col": "AF"},
        {"name": "Delivery of Major Materials","col": "AG"},
        {"name": "PIM",                        "col": "AH"},
        {"name": "Start Production",           "col": "AI"},
        {"name": "FAT",                        "col": "AJ"},
        {"name": "Ready for Shipment",         "col": "AK"},
        {"name": "Receipt at Worksite",        "col": "AL"},
        {"name": "Punch List Clearance",       "col": "AM"},
        {"name": "Final Documentation",        "col": "AN"},
    ],
}


def col_to_idx(col_str):
    """Convert Excel column letter(s) to 1-based index."""
    if not col_str:
        return None
    result = 0
    for char in col_str.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def normalize_header(text):
    """Normalize header text for comparison: lowercase, strip, collapse whitespace/newlines."""
    if text is None:
        return ""
    s = str(text).replace('\n', ' ').replace('\r', ' ')
    return ' '.join(s.strip().split()).lower()


def _match_stage_column(stage_name, col_map):
    """Find the column letter for a stage by matching its name against header texts.

    Matching priority:
      1. Exact match (after normalization)
      2. Containment: stage name is a substring of the header
      3. Word overlap: >= 3 shared words (handles abbreviation changes)
    Returns the column letter or None if no match is found.
    """
    norm = normalize_header(stage_name)

    # 1. Exact match
    if norm in col_map:
        return col_map[norm]

    # 1b. Check aliases for Z1F / ASK template differences
    ALIASES = {
        "mr ifa issued": ["mr issued (ifr)", "mr issued"],
        "vd submission": ["vdb submission"],
        "approval of key vd": ["approval of key vp"],
        "delivery of major materials": ["main material arrived"],
    }
    if norm in ALIASES:
        for alias in ALIASES[norm]:
            if alias in col_map:
                return col_map[alias]
            for header, col in col_map.items():
                if alias in header:
                    return col

    # 2. Containment: stage name found within header text
    for header, col in col_map.items():
        if norm in header:
            return col

    # 3. Word overlap (>= 3 shared words to avoid false positives like
    #    "MR IFA Issued" matching "MR Issued")
    stage_words = set(norm.split())
    best_col = None
    best_overlap = 0
    for header, col in col_map.items():
        header_words = set(header.split())
        overlap = len(stage_words & header_words)
        if overlap >= 3 and overlap > best_overlap:
            best_overlap = overlap
            best_col = col

    return best_col


def auto_detect_columns(ws, base_config):
    """Auto-detect column positions by scanning the header row.

    Handles column insertions, deletions, and renames between file revisions.
    Works reliably after Git clone or Render deploy where file metadata is lost.
    Falls back to the hardcoded base_config values if detection fails.
    """
    header_row = base_config["header_row"]

    # Build normalized-header → column-letter map (skip date-valued cells)
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None or isinstance(v, (datetime.datetime, datetime.date)):
            continue
        norm = normalize_header(v)
        if norm:
            col_map[norm] = get_column_letter(c)

    config = dict(base_config)

    # --- P/F/A column ---
    if "p/f/a" in col_map:
        config["pfa_col"] = col_map["p/f/a"]
    else:
        # P/F/A header may be missing (e.g. ASK) — scan first data rows for known values
        pfa_keys = set(base_config["pfa_values"].keys())
        for c in range(1, ws.max_column + 1):
            hits = sum(
                1 for r in range(base_config["data_start_row"],
                                 min(base_config["data_start_row"] + 9, ws.max_row + 1))
                if str(ws.cell(row=r, column=c).value or "").strip().upper() in pfa_keys
            )
            if hits >= 3:
                config["pfa_col"] = get_column_letter(c)
                break

    # --- Metadata columns ---
    META_PATTERNS = {
        "item_no_col":      ["no.", "no"],
        "package_name_col": ["package name"],
        "rfq_no_col":       ["rfq no.", "rfq no"],
        "mr_no_col":        ["mr no.", "mr no"],
        "priority_col":     ["priority level", "priority"],
        "lli_col":          ["lli (y/n)", "long lead item (y/n)", "lli", "long lead item"],
    }
    for key, patterns in META_PATTERNS.items():
        if base_config.get(key) is None:
            continue
        for p in patterns:
            if p in col_map:
                config[key] = col_map[p]
                break

    # --- Stage columns ---
    new_stages = []
    for stage in base_config["stages"]:
        if stage["col"] is None:
            new_stages.append({"name": stage["name"], "col": None})
        else:
            col = _match_stage_column(stage["name"], col_map)
            if col:
                new_stages.append({"name": stage["name"], "col": col})
            else:
                new_stages.append({"name": stage["name"], "col": stage["col"]})

    if new_stages:
        config["stages"] = new_stages

    # Log detected layout
    print(f"    [Auto-Detect] P/F/A col: {config['pfa_col']}  |  "
          f"Stages found: {len(config['stages'])}/{len(base_config['stages'])}")

    return config



def safe_date(value):
    """Extract a date from a cell value. Returns ISO string or None."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        if value.year < 2020:
            return None
        return value.strftime("%Y-%m-%d")
    return None


def get_file_sort_key(filepath):
    """Sort key for selecting the latest Excel file.

    Designed to work reliably across local dev, Git clones, and Render deploys
    where st_mtime is reset on checkout/deploy and cannot be trusted.

    Priority order (highest wins):
      1. has_date_suffix  — files with YYYYMMDD in the name always beat those without
      2. date_val         — the extracted YYYYMMDD integer (larger = newer)
      3. st_mtime         — tie-breaker for files with identical date suffix
      4. name length/name — deterministic last-resort tie-breaker
    """
    name = filepath.name
    # Try strict YYYYMMDD first, then flexible separators (2026-07-03, 2026_07_03, etc.)
    m = re.search(r'(20\d{6})', name)
    if not m:
        m = re.search(r'(20\d{2}[-_.]?\d{2}[-_.]?\d{2})', name)
    if m:
        has_date_suffix = 1          # rank above undated files
        date_str = re.sub(r'[^0-9]', '', m.group(0))
        date_val = int(date_str)
    else:
        has_date_suffix = 0          # undated file — only chosen when no dated file exists
        date_val = 0                 # don't use mtime as date — unreliable after git clone / deploy
    return (has_date_suffix, date_val, filepath.stat().st_mtime, len(name), name)


def find_excel_files():
    """Find the latest Z1F and ASK Excel files in the base directory based on date suffix and date modified."""
    z1f_candidates = []
    ask_candidates = []
    for f in BASE_DIR.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        fname = f.name.lower()
        if "z1f" in fname:
            z1f_candidates.append(f)
        elif "ask" in fname:
            ask_candidates.append(f)

    z1f_file = max(z1f_candidates, key=get_file_sort_key) if z1f_candidates else None
    ask_file = max(ask_candidates, key=get_file_sort_key) if ask_candidates else None
    
    if z1f_file:
        print(f"  [Auto-Detect] Selected latest Z1F file: {z1f_file.name}")
    if ask_file:
        print(f"  [Auto-Detect] Selected latest ASK file: {ask_file.name}")
        
    return z1f_file, ask_file


def extract_file_data(filepath, config):
    """Extract procurement data from a single Excel file."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    if config["sheet_name"] not in wb.sheetnames:
        # Try to find a matching sheet
        for sn in wb.sheetnames:
            if config["sheet_name"].lower() in sn.lower():
                config["sheet_name"] = sn
                break
        else:
            return {"error": f"Sheet '{config['sheet_name']}' not found. Available: {wb.sheetnames}"}

    ws = wb[config["sheet_name"]]

    # Auto-detect column layout from header row
    config = auto_detect_columns(ws, config)

    pfa_col_idx = col_to_idx(config["pfa_col"])
    pkg_col_idx = col_to_idx(config["package_name_col"])
    rfq_col_idx = col_to_idx(config["rfq_no_col"])
    item_col_idx = col_to_idx(config["item_no_col"])
    priority_col_idx = col_to_idx(config["priority_col"])
    lli_col_idx = col_to_idx(config["lli_col"])
    mr_col_idx = col_to_idx(config["mr_no_col"]) if config["mr_no_col"] else None

    stage_col_indices = [(s["name"], col_to_idx(s["col"])) for s in config["stages"]]

    packages = []
    row = config["data_start_row"]

    while row <= ws.max_row:
        # Check if this row is the start of a package (a Plan row)
        pfa_val = ws.cell(row=row, column=pfa_col_idx).value
        if pfa_val is None:
            row += 1
            continue

        pfa_str = str(pfa_val).strip().upper()
        mapped_val = config["pfa_values"].get(pfa_str)
        if mapped_val != "Plan":
            row += 1
            continue

        # Found the start of a package group at `row`
        plan_row = row
        
        # Scan forward to find where this package group ends (at the NEXT Plan row or end of sheet)
        group_rows = [plan_row]
        next_row = plan_row + 1
        while next_row <= ws.max_row:
            next_pfa_val = ws.cell(row=next_row, column=pfa_col_idx).value
            if next_pfa_val is not None:
                next_pfa_str = str(next_pfa_val).strip().upper()
                if config["pfa_values"].get(next_pfa_str) == "Plan":
                    # Reached the next package's Plan row!
                    break
            group_rows.append(next_row)
            next_row += 1

        # Identify forecast_row and actual_row within this group
        forecast_row = None
        actual_row = None
        for r in group_rows:
            val = ws.cell(row=r, column=pfa_col_idx).value
            if val is None:
                continue
            val_str = str(val).strip().upper()
            mapped = config["pfa_values"].get(val_str)
            if mapped == "Forecast":
                forecast_row = r
            elif mapped == "Actual":
                actual_row = r

        # Extract package info from the Plan row (or first available row with package name in group)
        info_row = plan_row
        package_name = ws.cell(row=info_row, column=pkg_col_idx).value
        if package_name is None:
            for r in group_rows:
                package_name = ws.cell(row=r, column=pkg_col_idx).value
                if package_name is not None:
                    info_row = r
                    break

        if package_name is None:
            row = next_row
            continue

        rfq_no = ws.cell(row=info_row, column=rfq_col_idx).value or ""
        item_no = ws.cell(row=info_row, column=item_col_idx).value or ""
        priority = ws.cell(row=info_row, column=priority_col_idx).value or ""
        lli = ws.cell(row=info_row, column=lli_col_idx).value or ""
        mr_no = ""
        if mr_col_idx:
            mr_no = ws.cell(row=info_row, column=mr_col_idx).value or ""

        # Extract stage dates
        stages = []
        for stage_name, stage_col in stage_col_indices:
            if stage_col:
                plan_date = safe_date(ws.cell(row=plan_row, column=stage_col).value) if plan_row else None
                forecast_date = safe_date(ws.cell(row=forecast_row, column=stage_col).value) if forecast_row else None
                actual_date = safe_date(ws.cell(row=actual_row, column=stage_col).value) if actual_row else None
            else:
                plan_date = None
                forecast_date = None
                actual_date = None

            stages.append({
                "name": stage_name,
                "plan": plan_date,
                "forecast": forecast_date,
                "actual": actual_date,
            })

        packages.append({
            "item_no": str(item_no),
            "package_name": str(package_name).strip(),
            "rfq_no": str(rfq_no).strip(),
            "mr_no": str(mr_no).strip(),
            "priority": str(priority).strip(),
            "lli": str(lli).strip(),
            "stages": stages,
        })

        # Move to the next package group
        row = next_row

    wb.close()
    return packages


def extract_all_data():
    """Extract data from both Excel files and compute delays."""
    z1f_file, ask_file = find_excel_files()

    today = datetime.date.today().isoformat()
    lookahead_end = (datetime.date.today() + datetime.timedelta(days=7)).isoformat()

    result = {
        "generated_at": datetime.datetime.now().isoformat(),
        "today": today,
        "lookahead_end": lookahead_end,
        "projects": {}
    }

    if z1f_file:
        print(f"  Reading Z1F: {z1f_file.name}")
        z1f_data = extract_file_data(z1f_file, Z1F_CONFIG.copy())
        if isinstance(z1f_data, dict) and "error" in z1f_data:
            result["projects"]["Z1F"] = z1f_data
        else:
            result["projects"]["Z1F"] = {
                "filename": z1f_file.name,
                "package_count": len(z1f_data),
                "packages": z1f_data,
            }
    else:
        result["projects"]["Z1F"] = {"error": "No Z1F Excel file found"}

    if ask_file:
        print(f"  Reading ASK: {ask_file.name}")
        ask_data = extract_file_data(ask_file, ASK_CONFIG.copy())
        if isinstance(ask_data, dict) and "error" in ask_data:
            result["projects"]["ASK"] = ask_data
        else:
            result["projects"]["ASK"] = {
                "filename": ask_file.name,
                "package_count": len(ask_data),
                "packages": ask_data,
            }
    else:
        result["projects"]["ASK"] = {"error": "No ASK Excel file found"}

    return result


# ─── HTTP Server ─────────────────────────────────────────────────────────────

# Data cache to avoid re-reading on every request
_data_cache = {"data": None, "lock": threading.Lock()}


class DashboardHandler(http.server.SimpleHTTPRequestHandler):
    """Custom HTTP handler for serving dashboard and API endpoints."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def do_GET(self):
        if self.path == "/api/data":
            self.send_api_data()
        elif self.path == "/api/refresh":
            self.send_api_refresh()
        elif self.path == "/":
            self.path = "/index.html"
            super().do_GET()
        elif self.path == "/favicon.ico":
            self.send_response(204)
            self.end_headers()
        else:
            super().do_GET()

    def do_POST(self):
        if self.path.startswith("/api/upload"):
            self.send_api_upload()
        else:
            self.send_error(404, "Not Found")

    def send_api_upload(self):
        """Handle weekly Excel file upload with dynamic YYYYMMDD naming."""
        try:
            content_length = int(self.headers.get("Content-Length", 0))
            if content_length <= 0 or content_length > 50 * 1024 * 1024:
                raise ValueError("Invalid content length (max 50MB)")

            body = self.rfile.read(content_length)
            content_type = self.headers.get("Content-Type", "")

            project_type = ""
            orig_filename = ""
            file_bytes = None

            if "multipart/form-data" in content_type:
                match = re.search(r'boundary=([^\s;]+)', content_type, re.IGNORECASE)
                if not match:
                    raise ValueError("Missing boundary in Content-Type")
                boundary = match.group(1).strip('"')
                boundary_bytes = ("--" + boundary).encode("utf-8")

                parts = body.split(boundary_bytes)
                for part in parts:
                    if not part or part == b'--\r\n' or part == b'--\n' or part == b'--':
                        continue
                    if b'\r\n\r\n' in part:
                        header_bytes, content = part.split(b'\r\n\r\n', 1)
                        if content.endswith(b'\r\n'):
                            content = content[:-2]
                        elif content.endswith(b'\n'):
                            content = content[:-1]
                    elif b'\n\n' in part:
                        header_bytes, content = part.split(b'\n\n', 1)
                        if content.endswith(b'\r\n'):
                            content = content[:-2]
                        elif content.endswith(b'\n'):
                            content = content[:-1]
                    else:
                        continue

                    header_str = header_bytes.decode("utf-8", errors="ignore")
                    if 'name="project"' in header_str or 'name="project_type"' in header_str:
                        project_type = content.decode("utf-8", errors="ignore").strip().upper()
                    elif 'filename=' in header_str or 'name="file"' in header_str:
                        file_bytes = content
                        fn_match = re.search(r'filename="([^"]+)"', header_str)
                        if not fn_match:
                            fn_match = re.search(r'filename=([^\s;]+)', header_str)
                        if fn_match:
                            orig_filename = fn_match.group(1).strip()
            else:
                file_bytes = body
                from urllib.parse import urlparse, parse_qs
                qs = parse_qs(urlparse(self.path).query)
                project_type = qs.get("project", [""])[0].upper()
                orig_filename = qs.get("filename", ["uploaded.xlsx"])[0]

            if not file_bytes:
                raise ValueError("No file content received")

            if not project_type:
                if "Z1F" in orig_filename.upper():
                    project_type = "Z1F"
                elif "ASK" in orig_filename.upper():
                    project_type = "ASK"

            if project_type not in ("Z1F", "ASK"):
                raise ValueError("Could not determine project (Z1F or ASK) from upload")

            # Extract date from orig_filename or use current date YYYYMMDD
            date_str = datetime.datetime.now().strftime("%Y%m%d")
            m = re.search(r'(20\d{6})', orig_filename)
            if not m:
                m = re.search(r'(20\d{2}[-_.]?\d{2}[-_.]?\d{2})', orig_filename)
            if m:
                date_str = re.sub(r'[^0-9]', '', m.group(0))

            if project_type == "Z1F":
                target_filename = f"Attachment 1-Procurement Plan-Z1F - {date_str}.xlsx"
            else:
                target_filename = f"Attachment 2-Procurement Plan-ASK - {date_str}.xlsx"

            target_path = BASE_DIR / target_filename
            with open(target_path, "wb") as f:
                f.write(file_bytes)

            print(f"\n[API] Uploaded and saved {project_type} file: {target_filename} ({len(file_bytes)} bytes)")

            print("[API] Refreshing data cache after upload...")
            new_data = extract_all_data()
            with _data_cache["lock"]:
                _data_cache["data"] = new_data

            response_data = {
                "success": True,
                "project": project_type,
                "filename": target_filename,
                "data": new_data
            }
            json_bytes = json.dumps(response_data, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", len(json_bytes))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json_bytes)
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = json.dumps({"error": str(e)}).encode("utf-8")
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(error_msg))
            self.end_headers()
            self.wfile.write(error_msg)

    def send_api_data(self):
        """Return cached data, or read Excel files if cache is empty."""
        try:
            with _data_cache["lock"]:
                if _data_cache["data"] is None:
                    print("\n[API] First load - extracting data from Excel files...")
                    _data_cache["data"] = extract_all_data()
                data = _data_cache["data"]

            json_bytes = json.dumps(data, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", len(json_bytes))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json_bytes)
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = json.dumps({"error": str(e)}).encode("utf-8")
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(error_msg))
            self.end_headers()
            self.wfile.write(error_msg)

    def send_api_refresh(self):
        """Force re-read Excel files (called by Update button)."""
        try:
            print("\n[API] Refreshing data from Excel files...")
            new_data = extract_all_data()
            with _data_cache["lock"]:
                _data_cache["data"] = new_data
            print(f"[API] Refresh complete. Z1F: {new_data['projects'].get('Z1F', {}).get('package_count', '?')} pkgs, ASK: {new_data['projects'].get('ASK', {}).get('package_count', '?')} pkgs")

            json_bytes = json.dumps(new_data, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", len(json_bytes))
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json_bytes)
        except Exception as e:
            import traceback
            traceback.print_exc()
            error_msg = json.dumps({"error": str(e)}).encode("utf-8")
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(error_msg))
            self.end_headers()
            self.wfile.write(error_msg)

    def log_message(self, format, *args):
        # Suppress default access logs for static files, only log API calls
        try:
            msg = str(args[0]) if args else ""
            if "/api/" in msg:
                super().log_message(format, *args)
        except Exception:
            pass


class ThreadedHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    """Handle requests in separate threads."""
    daemon_threads = True


# ─── Main ────────────────────────────────────────────────────────────────────

def preload_data():
    print("\n  [Background] Pre-loading Excel data...")
    try:
        with _data_cache["lock"]:
            if _data_cache["data"] is None:
                _data_cache["data"] = extract_all_data()
        z1f_count = _data_cache["data"]["projects"].get("Z1F", {}).get("package_count", "?")
        ask_count = _data_cache["data"]["projects"].get("ASK", {}).get("package_count", "?")
        print(f"  [Background] Loaded: Z1F={z1f_count} packages, ASK={ask_count} packages\n")
    except Exception as e:
        print(f"  [Background] Warning: Could not pre-load data: {e}\n")


if __name__ == "__main__":
    print("=" * 54)
    print("   Procurement Tracking Dashboard Server")
    print("=" * 54)
    print(f"   URL: http://localhost:{PORT}")
    print("   Press Ctrl+C to stop")
    print("=" * 54)

    z1f, ask = find_excel_files()
    print(f"\n  Z1F file: {z1f.name if z1f else 'NOT FOUND'}")
    print(f"  ASK file: {ask.name if ask else 'NOT FOUND'}")
    print(f"\n  Drop new Excel files in: {BASE_DIR}")
    print("  Dashboard will re-read files on Update click.")

    # Start pre-loading in background thread
    threading.Thread(target=preload_data, daemon=True).start()

    print(f"\n  Server ready at http://localhost:{PORT}\n")

    server = ThreadedHTTPServer(("", PORT), DashboardHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\nServer stopped.")
        server.server_close()

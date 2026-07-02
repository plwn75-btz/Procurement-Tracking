import openpyxl
from openpyxl.utils import get_column_letter

# === Z1F File ===
print("=" * 80)
print("Z1F FILE - PSR Sheet")
print("=" * 80)
wb = openpyxl.load_workbook(r'Attachment 1-Procurement Plan-Z1F.xlsx', data_only=True)
ws = wb['PSR']

print(f"\nDimensions: {ws.dimensions}")
print(f"Rows: {ws.min_row}-{ws.max_row}, Cols: {ws.min_column}-{ws.max_column}")

print("\n--- Header area (rows 1-13) ---")
for row_idx in range(1, 14):
    vals = []
    for col_idx in range(1, 60):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            vals.append(f"{col_letter}={repr(cell.value)}")
    if vals:
        print(f"  Row {row_idx}: " + "  |  ".join(vals))

print("\n--- Merged cells (sample) ---")
for i, mc in enumerate(ws.merged_cells.ranges):
    if i < 30:
        print(f"  {mc}")

# === ASK File ===
print("\n" + "=" * 80)
print("ASK FILE - PSR Overall Cycle Sheet")
print("=" * 80)
wb2 = openpyxl.load_workbook(r'Attachment 2-Procurement Plan-ASK.xlsx', data_only=True)
ws2 = wb2['PSR Overall Cycle']

print(f"\nDimensions: {ws2.dimensions}")
print(f"Rows: {ws2.min_row}-{ws2.max_row}, Cols: {ws2.min_column}-{ws2.max_column}")

print("\n--- Header area (rows 1-18) ---")
for row_idx in range(1, 19):
    vals = []
    for col_idx in range(1, 50):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            vals.append(f"{col_letter}={repr(cell.value)}")
    if vals:
        print(f"  Row {row_idx}: " + "  |  ".join(vals))

print("\n--- Merged cells (sample) ---")
for i, mc in enumerate(ws2.merged_cells.ranges):
    if i < 40:
        print(f"  {mc}")

# Now get first data rows from ASK
print("\n--- ASK Data rows 16-25 ---")
for row_idx in range(16, 26):
    vals = []
    for col_idx in range(1, 48):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            vals.append(f"{col_letter}={repr(cell.value)}")
    if vals:
        print(f"  Row {row_idx}: " + "  |  ".join(vals))
    else:
        print(f"  Row {row_idx}: (empty)")

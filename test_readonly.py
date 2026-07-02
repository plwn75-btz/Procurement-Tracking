"""Quick test to verify read_only mode works with the extraction logic."""
import openpyxl
import datetime

# Test Z1F
print("Testing Z1F file with read_only=True...")
wb = openpyxl.load_workbook(r'Attachment 1-Procurement Plan-Z1F.xlsx', data_only=True, read_only=True)
ws = wb['PSR']
print(f"  Sheet: {ws.title}")

# In read_only mode, we need to iterate rows differently
# Let's test cell access
try:
    val = ws.cell(row=10, column=4).value
    print(f"  Cell D10 (direct): {val}")
except Exception as e:
    print(f"  Direct cell access error: {e}")

# Test iterating rows
print("  Testing row iteration...")
for row in ws.iter_rows(min_row=10, max_row=10, min_col=1, max_col=15, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f"    {cell.coordinate} = {repr(cell.value)[:50]}")

# Test data row
print("  Testing data row 13...")
for row in ws.iter_rows(min_row=13, max_row=13, min_col=1, max_col=15, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f"    {cell.coordinate} = {repr(cell.value)[:50]}")

wb.close()
print("\nDone!")

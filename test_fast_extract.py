import time
import openpyxl
from pathlib import Path

t0 = time.time()
print("Loading Z1F workbook (standard mode)...")
wb = openpyxl.load_workbook('Attachment 1-Procurement Plan-Z1F.xlsx', data_only=True)
t1 = time.time()
print(f"Loaded Z1F in {t1-t0:.2f}s. max_row={wb['PSR'].max_row}")

print("Loading ASK workbook (standard mode)...")
wb2 = openpyxl.load_workbook('Attachment 2-Procurement Plan-ASK.xlsx', data_only=True)
t2 = time.time()
print(f"Loaded ASK in {t2-t1:.2f}s. max_row={wb2['PSR Overall Cycle'].max_row}")

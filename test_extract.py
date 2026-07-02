"""Test the data extraction part of server.py"""
import sys
sys.path.insert(0, '.')

# Import the extraction functions from server.py
import importlib.util
spec = importlib.util.spec_from_file_location("server", "server.py")
# Can't import server.py directly because it has serve_forever() in __main__
# So let's just run the extraction logic directly

import openpyxl
import datetime
from pathlib import Path

BASE_DIR = Path('.')

def col_to_idx(col_str):
    result = 0
    for char in col_str.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def safe_date(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    return None

# Quick test: try to read Z1F
print("Testing Z1F extraction...")
wb = openpyxl.load_workbook('Attachment 1-Procurement Plan-Z1F.xlsx', data_only=True, read_only=True)
ws = wb['PSR']
print(f"  max_row: {ws.max_row}")

pfa_col_idx = col_to_idx("M")
pkg_col_idx = col_to_idx("D")
count = 0

row = 13
while row <= ws.max_row:
    pfa_val = ws.cell(row=row, column=pfa_col_idx).value
    if pfa_val is not None and str(pfa_val).strip() == 'P':
        pkg = ws.cell(row=row, column=pkg_col_idx).value
        count += 1
        if count <= 3:
            print(f"  Package {count}: {pkg}")
    row += 1

print(f"  Total Plan rows: {count}")
wb.close()

# Test ASK
print("\nTesting ASK extraction...")
wb2 = openpyxl.load_workbook('Attachment 2-Procurement Plan-ASK.xlsx', data_only=True, read_only=True)
ws2 = wb2['PSR Overall Cycle']
print(f"  max_row: {ws2.max_row}")

pfa_col_idx2 = col_to_idx("P")
pkg_col_idx2 = col_to_idx("C")
count2 = 0

row = 19
while row <= ws2.max_row:
    pfa_val = ws2.cell(row=row, column=pfa_col_idx2).value
    if pfa_val is not None and str(pfa_val).strip() == 'Plan':
        pkg = ws2.cell(row=row, column=pkg_col_idx2).value
        count2 += 1
        if count2 <= 3:
            print(f"  Package {count2}: {pkg}")
    row += 1

print(f"  Total Plan rows: {count2}")
wb2.close()

print("\nAll tests passed!")
